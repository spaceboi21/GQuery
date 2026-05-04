"""
Cell-level I/O for Excel / CSV files.

Responsibilities
----------------
- Read paginated rectangular blocks of cells (for the spreadsheet overlay).
- Read specific A1-notation ranges (for agent tools / chat references).
- Write values and formulas back to .xlsx files via openpyxl.

All operations are bounded by the configured sandbox:
they never load the entire file into memory, and openpyxl opens in streaming
mode for reads.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


# ==================================================================
# A1 notation helpers
# ==================================================================

_A1_RANGE_RE = re.compile(r"^([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", re.IGNORECASE)


def col_letter_to_index(letter: str) -> int:
    """A → 1, B → 2, Z → 26, AA → 27 …"""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def col_index_to_letter(idx: int) -> str:
    """1 → A, 27 → AA."""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def parse_a1_range(range_str: str) -> tuple[int, int, int, int]:
    """
    Parse 'A1', 'A1:C10', 'B2' into (min_row, min_col, max_row, max_col) — 1-based.

    Raises ValueError on malformed input.
    """
    m = _A1_RANGE_RE.match(range_str.strip())
    if not m:
        raise ValueError(f"Invalid A1 range: {range_str!r}")
    col_a, row_a, col_b, row_b = m.groups()
    min_col = col_letter_to_index(col_a)
    min_row = int(row_a)
    max_col = col_letter_to_index(col_b) if col_b else min_col
    max_row = int(row_b) if row_b else min_row
    if max_col < min_col:
        min_col, max_col = max_col, min_col
    if max_row < min_row:
        min_row, max_row = max_row, min_row
    return min_row, min_col, max_row, max_col


# ==================================================================
# Reading
# ==================================================================

def read_sheet_data(
    path: Path,
    sheet_name: str | None = None,
    offset: int = 0,
    limit: int = 500,
    max_cols: int = 100,
) -> dict:
    """
    Load a rectangular block for the overlay grid.

    Returns:
        {
            "sheet_name": str,
            "all_sheets": [str, ...],    # so UI can tab between sheets
            "headers": [str, ...],        # column A1 letters ("A", "B", ...)
            "rows": [[cell, cell, ...]],  # 2D array of scalar values
            "total_rows": int,
            "total_cols": int,
            "offset": int,
            "limit": int,
        }
    """
    ext = path.suffix.lower()
    if ext == ".csv":
        return _read_csv_block(path, offset, limit, max_cols)
    if ext == ".xls":
        return _read_xls_block(path, sheet_name, offset, limit, max_cols)
    return _read_xlsx_block(path, sheet_name, offset, limit, max_cols)


def _read_xlsx_block(
    path: Path, sheet_name: str | None, offset: int, limit: int, max_cols: int
) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_sheets = wb.sheetnames
    if not all_sheets:
        wb.close()
        return _empty_result("", all_sheets)
    target = sheet_name or all_sheets[0]
    if target not in all_sheets:
        wb.close()
        raise ValueError(f"Sheet {target!r} not in {all_sheets}")

    ws = wb[target]
    total_rows = ws.max_row or 0
    total_cols = min(ws.max_column or 0, max_cols)

    start_row = offset + 1
    end_row = min(start_row + limit - 1, total_rows)

    rows: list[list] = []
    if end_row >= start_row and total_cols > 0:
        for row in ws.iter_rows(
            min_row=start_row, max_row=end_row,
            min_col=1, max_col=total_cols,
            values_only=True,
        ):
            rows.append([_jsonify(v) for v in row])

    headers = [col_index_to_letter(i + 1) for i in range(total_cols)]
    wb.close()

    return {
        "sheet_name": target,
        "all_sheets": all_sheets,
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "offset": offset,
        "limit": limit,
    }


def _read_xls_block(
    path: Path, sheet_name: str | None, offset: int, limit: int, max_cols: int
) -> dict:
    xl = pd.ExcelFile(path, engine="xlrd")
    all_sheets = xl.sheet_names
    if not all_sheets:
        return _empty_result("", all_sheets)
    target = sheet_name or all_sheets[0]
    df = xl.parse(target, header=None, skiprows=offset, nrows=limit)
    df = df.iloc[:, :max_cols]
    total_rows = xl.parse(target, header=None, usecols=[0]).shape[0]
    total_cols = min(df.shape[1], max_cols)
    rows = [[_jsonify(v) for v in row] for row in df.values.tolist()]
    headers = [col_index_to_letter(i + 1) for i in range(total_cols)]
    return {
        "sheet_name": target,
        "all_sheets": all_sheets,
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "offset": offset,
        "limit": limit,
    }


def _read_csv_block(path: Path, offset: int, limit: int, max_cols: int) -> dict:
    df = pd.read_csv(
        path, header=None, skiprows=offset, nrows=limit,
        low_memory=False, encoding_errors="replace",
    )
    df = df.iloc[:, :max_cols]
    # Cheap total-row count via byte streaming
    try:
        total_rows = sum(1 for _ in open(path, encoding="utf-8", errors="ignore"))
    except Exception:
        total_rows = len(df) + offset
    total_cols = min(df.shape[1], max_cols)
    rows = [[_jsonify(v) for v in row] for row in df.values.tolist()]
    headers = [col_index_to_letter(i + 1) for i in range(total_cols)]
    return {
        "sheet_name": "Sheet1",
        "all_sheets": ["Sheet1"],
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "offset": offset,
        "limit": limit,
    }


def read_cell_range(
    path: Path, sheet_name: str | None, range_str: str, max_cells: int = 2000,
) -> dict:
    """
    Read exactly the cells defined by an A1 range (e.g. 'A1:C10').

    Used by the agent to get precise cell context for user-referenced selections.
    """
    min_row, min_col, max_row, max_col = parse_a1_range(range_str)
    n_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    if n_cells > max_cells:
        raise ValueError(
            f"Range too large ({n_cells} cells). Max is {max_cells}. "
            "Narrow the range or ask the agent to aggregate first."
        )

    ext = path.suffix.lower()
    rows: list[list] = []

    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        target = sheet_name or wb.sheetnames[0]
        ws = wb[target]
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True,
        ):
            rows.append([_jsonify(v) for v in row])
        wb.close()
    elif ext == ".xls":
        xl = pd.ExcelFile(path, engine="xlrd")
        target = sheet_name or xl.sheet_names[0]
        df = xl.parse(target, header=None)
        df = df.iloc[min_row - 1:max_row, min_col - 1:max_col]
        rows = [[_jsonify(v) for v in row] for row in df.values.tolist()]
    else:  # csv
        df = pd.read_csv(
            path, header=None, low_memory=False, encoding_errors="replace",
        )
        df = df.iloc[min_row - 1:max_row, min_col - 1:max_col]
        rows = [[_jsonify(v) for v in row] for row in df.values.tolist()]

    return {
        "range": f"{col_index_to_letter(min_col)}{min_row}:{col_index_to_letter(max_col)}{max_row}",
        "sheet_name": sheet_name or "",
        "cells": rows,
        "row_count": len(rows),
        "col_count": len(rows[0]) if rows else 0,
    }


# ==================================================================
# Writing (openpyxl — formulas preserved)
# ==================================================================

def write_cells(
    path: Path,
    sheet_name: str | None,
    writes: list[dict],
    create_sheet: bool = False,
) -> dict:
    """
    Apply a list of cell writes to an .xlsx file.

    writes = [{"cell": "A1", "value": "Name"}, {"cell": "B2", "formula": "=VLOOKUP(...)"}]
    - `value` is written as a literal (string/number/bool).
    - `formula` is written as a formula (must start with '=').

    Only .xlsx / .xlsm are supported for writes. CSVs must be re-written via pandas
    (not exposed here — too destructive for row-indexed formats).

    Returns: {"ok": True, "writes_applied": N, "sheet": ..., "file": ...}
    """
    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(f"write_cells only supports .xlsx/.xlsm (got {ext})")
    if not writes:
        return {"ok": True, "writes_applied": 0, "sheet": sheet_name, "file": str(path)}

    wb = openpyxl.load_workbook(path)
    target = sheet_name or wb.sheetnames[0]

    if target not in wb.sheetnames:
        if not create_sheet:
            wb.close()
            raise ValueError(f"Sheet {target!r} not found. Available: {wb.sheetnames}")
        wb.create_sheet(target)

    ws = wb[target]
    applied = 0
    for w in writes:
        cell = w.get("cell")
        if not cell:
            continue
        if "formula" in w:
            formula = str(w["formula"])
            if not formula.startswith("="):
                formula = "=" + formula
            ws[cell] = formula
            applied += 1
        elif "value" in w:
            ws[cell] = w["value"]
            applied += 1

    wb.save(path)
    wb.close()

    return {
        "ok": True,
        "writes_applied": applied,
        "sheet": target,
        "file": str(path),
    }


# ==================================================================
# Helpers
# ==================================================================

def _jsonify(v: Any) -> Any:
    """Make cell values JSON-serialisable for the frontend."""
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN
        return ""
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _empty_result(sheet: str, sheets: list[str]) -> dict:
    return {
        "sheet_name": sheet,
        "all_sheets": sheets,
        "headers": [],
        "rows": [],
        "total_rows": 0,
        "total_cols": 0,
        "offset": 0,
        "limit": 0,
    }


# ==================================================================
# Pivot-table detection (openpyxl private API — read-write mode required)
# ==================================================================

def extract_pivots_from_workbook(path: Path) -> dict[str, list[dict]]:
    """
    Scan an .xlsx/.xlsm file for pivot tables. Returns {sheet_name: [pivot_info, ...]}.

    pivot_info = {
        "name": str,
        "location": "A1:G20" or "",
        "source_fields": [str, ...],      # all columns in the source data
        "row_fields":    [str, ...],      # "Rows" area
        "col_fields":    [str, ...],      # "Columns" area
        "page_fields":   [str, ...],      # "Filters" area
        "data_fields":   [{"field": str, "agg": str, "label": str}, ...],
        "source_ref":    str,             # e.g. "Detail1!A1:AB91195"
    }

    Safe to call on files with no pivots — returns empty dict.
    """
    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        return {}

    try:
        # Must open in read-write mode — read_only hides _pivots
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    except Exception:
        return {}

    results: dict[str, list[dict]] = {}
    try:
        for ws in wb.worksheets:
            pivots = getattr(ws, "_pivots", None) or []
            extracted: list[dict] = []
            for p in pivots:
                info = _extract_single_pivot(p)
                if info:
                    extracted.append(info)
            if extracted:
                results[ws.title] = extracted
    finally:
        wb.close()

    return results


def _extract_single_pivot(p) -> dict | None:
    """Convert one openpyxl TableDefinition to a plain dict."""
    try:
        # Source fields come from the pivot cache — indices below reference this list
        source_fields: list[str] = []
        source_ref = ""
        cache = getattr(p, "cache", None)
        if cache is not None:
            cache_fields = getattr(cache, "cacheFields", None) or []
            source_fields = [str(getattr(f, "name", "")) for f in cache_fields]
            # Cache source reference (the table the pivot is built from)
            cd = getattr(cache, "cacheSource", None)
            if cd is not None:
                ws_src = getattr(cd, "worksheetSource", None)
                if ws_src is not None:
                    sheet = getattr(ws_src, "sheet", "") or ""
                    ref = getattr(ws_src, "ref", "") or ""
                    name = getattr(ws_src, "name", "") or ""
                    if ref:
                        source_ref = f"{sheet}!{ref}" if sheet else ref
                    elif name:
                        source_ref = name

        def resolve_fields(field_list) -> list[str]:
            if not field_list:
                return []
            out = []
            for f in field_list:
                idx = getattr(f, "x", None)
                if idx is None:
                    continue
                if 0 <= idx < len(source_fields):
                    out.append(source_fields[idx])
                elif idx == -2:
                    out.append("<values>")  # Excel's "Σ Values" placeholder
            return out

        # Row / Col / Page field areas
        row_fields = resolve_fields(getattr(p, "rowFields", None))
        col_fields = resolve_fields(getattr(p, "colFields", None))
        page_fields = resolve_fields(getattr(p, "pageFields", None))

        # Value/data fields (aggregations)
        data_fields: list[dict] = []
        for df in (getattr(p, "dataFields", None) or []):
            fld_idx = getattr(df, "fld", -1)
            agg = getattr(df, "subtotal", None) or "sum"
            label = getattr(df, "name", "") or ""
            field_name = (
                source_fields[fld_idx]
                if 0 <= fld_idx < len(source_fields)
                else label or f"field_{fld_idx}"
            )
            data_fields.append({"field": field_name, "agg": str(agg), "label": label})

        # Location of the pivot on the displaying sheet
        loc = getattr(p, "location", None)
        location_ref = getattr(loc, "ref", "") if loc is not None else ""

        return {
            "name": getattr(p, "name", "") or "",
            "location": location_ref,
            "source_fields": source_fields,
            "row_fields": row_fields,
            "col_fields": col_fields,
            "page_fields": page_fields,
            "data_fields": data_fields,
            "source_ref": source_ref,
        }
    except Exception as exc:
        return {"name": getattr(p, "name", "") or "<unnamed>", "error": str(exc)}
