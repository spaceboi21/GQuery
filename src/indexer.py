"""
FileIndexer – crawl a directory of .xlsx/.xlsm/.xls/.csv files, extract metadata only
(never reads all cell data into memory), and persist to SQLite with FTS5.

Large-file strategy
-------------------
< threshold_mb  → pandas (fast, convenient)
>= threshold_mb → polars lazy scan (zero-copy schema, no full load)
                  openpyxl read_only for xlsx (streams row-by-row)
"""

from __future__ import annotations

import json
import os
import sqlite3
import time
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import polars as pl

from .config import Settings
from .graph import KnowledgeGraph
from .cells import extract_pivots_from_workbook
from .schema import ColumnMeta, FileMeta, IndexStats, SheetMeta

_DDL = """
CREATE TABLE IF NOT EXISTS files (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    path         TEXT UNIQUE NOT NULL,
    name         TEXT NOT NULL,
    file_type    TEXT NOT NULL,
    size_bytes   INTEGER,
    modified_at  REAL,
    indexed_at   REAL NOT NULL
);

CREATE TABLE IF NOT EXISTS sheets (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    file_id      INTEGER NOT NULL REFERENCES files(id) ON DELETE CASCADE,
    sheet_name   TEXT NOT NULL,
    row_count    INTEGER,
    col_count    INTEGER,
    headers_json TEXT NOT NULL,
    dtypes_json  TEXT NOT NULL,
    sample_json  TEXT NOT NULL DEFAULT '{}',
    UNIQUE(file_id, sheet_name)
);

CREATE VIRTUAL TABLE IF NOT EXISTS headers_fts USING fts5(
    header_name,
    sheet_id    UNINDEXED,
    file_path   UNINDEXED,
    tokenize    = 'porter ascii'
);

CREATE TABLE IF NOT EXISTS graph_state (
    id          INTEGER PRIMARY KEY CHECK (id = 1),
    nodes_json  TEXT NOT NULL DEFAULT '{}',
    edges_json  TEXT NOT NULL DEFAULT '[]',
    updated_at  REAL NOT NULL
);
"""


class FileIndexer:
    def __init__(self, settings: Settings) -> None:
        self.cfg = settings
        self.db = self._open_db()
        self.kg = KnowledgeGraph()
        if not self.kg.load(self.db):
            self._rebuild_kg()

    # ------------------------------------------------------------------
    # Public
    # ------------------------------------------------------------------

    def index_directory(self, force: bool = False) -> dict:
        """Crawl data_path, (re)index changed files, rebuild KG edges.

        force=True clears all existing records first and re-indexes every file.
        """
        if force:
            self.db.execute("DELETE FROM files")
            self.db.commit()

        added = updated = skipped = 0
        for fpath in self._iter_data_files():
            stat = fpath.stat()
            existing = self.db.execute(
                "SELECT id, modified_at FROM files WHERE path = ?", (str(fpath),)
            ).fetchone()
            if existing and existing[1] == stat.st_mtime and not force:
                skipped += 1
                continue
            try:
                meta = self._extract_meta(fpath)
                if existing:
                    self._delete_file(existing[0])
                    updated += 1
                else:
                    added += 1
                self._persist_file(meta)
            except Exception as exc:
                print(f"[indexer] skip {fpath.name}: {exc}")

        kg_edges = self._rebuild_kg()
        self.kg.save(self.db)
        return {"added": added, "updated": updated, "skipped": skipped, "kg_edges": kg_edges}

    def search_files(self, query: str = "", file_type: str = "") -> list[dict]:
        """Search indexed files. If query is provided, uses FTS5 on headers."""
        if query:
            hits = self.fts_search(query)
            file_paths = list(dict.fromkeys(h["file_path"] for h in hits))
        else:
            rows = self.db.execute(
                "SELECT path FROM files" + (" WHERE file_type = ?" if file_type else ""),
                (file_type,) if file_type else (),
            ).fetchall()
            file_paths = [r[0] for r in rows]

        results = []
        for fp in file_paths:
            row = self.db.execute(
                "SELECT id, name, file_type, size_bytes, modified_at FROM files WHERE path = ?",
                (fp,),
            ).fetchone()
            if not row:
                continue
            sheets = self.db.execute(
                "SELECT sheet_name, row_count, col_count, headers_json FROM sheets WHERE file_id = ?",
                (row[0],),
            ).fetchall()
            results.append({
                "path": fp,
                "name": row[1],
                "file_type": row[2],
                "size_mb": round((row[3] or 0) / 1_048_576, 2),
                "sheets": [
                    {
                        "sheet_name": s[0],
                        "row_count": s[1],
                        "col_count": s[2],
                        "headers": json.loads(s[3]),
                    }
                    for s in sheets
                ],
            })
        return results

    def get_schema(self, path: Path) -> FileMeta:
        """Return full FileMeta for a single file."""
        row = self.db.execute(
            "SELECT id, name, file_type, size_bytes, modified_at FROM files WHERE path = ?",
            (str(path),),
        ).fetchone()
        if not row:
            raise FileNotFoundError(f"{path} not indexed. Run index_directory() first.")

        sheets_rows = self.db.execute(
            "SELECT sheet_name, row_count, col_count, headers_json, dtypes_json, sample_json, "
            "header_row, sheet_type, pivot_json "
            "FROM sheets WHERE file_id = ?",
            (row[0],),
        ).fetchall()

        sheets = []
        for sr in sheets_rows:
            headers = json.loads(sr[3])
            dtypes = json.loads(sr[4])
            samples = json.loads(sr[5])
            header_row = sr[6] if sr[6] is not None else 0
            sheet_type = sr[7] if sr[7] else "data"
            pivot_tables = json.loads(sr[8]) if sr[8] else []
            cols = [
                ColumnMeta(
                    name=h,
                    dtype=dtypes.get(h, "unknown"),
                    sample_values=samples.get(h, []),
                )
                for h in headers
            ]
            sheets.append(SheetMeta(
                sheet_name=sr[0],
                row_count=sr[1] or 0,
                col_count=sr[2] or 0,
                columns=cols,
                header_row=header_row,
                sheet_type=sheet_type,
                pivot_tables=pivot_tables,
            ))

        return FileMeta(
            path=path,
            name=row[1],
            file_type=row[2],
            size_mb=round((row[3] or 0) / 1_048_576, 2),
            modified_at=row[4] or 0.0,
            sheets=sheets,
        )

    def fts_search(self, query: str, limit: int | None = None) -> list[dict]:
        """FTS5 header search. Returns list of {sheet_id, header, file_path}."""
        limit = limit or self.cfg.fts_result_limit
        safe_query = " OR ".join(
            f'"{term.strip()}"' for term in query.split() if term.strip()
        )
        if not safe_query:
            return []
        try:
            rows = self.db.execute(
                "SELECT sheet_id, header_name, file_path FROM headers_fts "
                "WHERE headers_fts MATCH ? LIMIT ?",
                (safe_query, limit),
            ).fetchall()
        except sqlite3.OperationalError:
            return []
        return [{"sheet_id": r[0], "header": r[1], "file_path": r[2]} for r in rows]

    def stats(self) -> IndexStats:
        total_files = self.db.execute("SELECT COUNT(*) FROM files").fetchone()[0]
        total_sheets = self.db.execute("SELECT COUNT(*) FROM sheets").fetchone()[0]
        total_columns = self.db.execute(
            "SELECT SUM(col_count) FROM sheets"
        ).fetchone()[0] or 0
        return IndexStats(
            total_files=total_files,
            total_sheets=total_sheets,
            total_columns=total_columns,
            total_edges=self.kg.edge_count(),
            data_path=str(self.cfg.data_path),
            db_path=str(self.cfg.db_path),
        )

    # ------------------------------------------------------------------
    # Internal – extraction
    # ------------------------------------------------------------------

    def _extract_meta(self, path: Path) -> FileMeta:
        stat = path.stat()
        size_mb = stat.st_size / 1_048_576
        is_large = size_mb >= self.cfg.large_file_threshold_mb
        ext = path.suffix.lower()

        if ext == ".csv":
            sheets = [self._extract_csv_sheet(path, is_large)]
        elif ext == ".xls":
            sheets = self._extract_xls_sheets(path)
        else:  # .xlsx, .xlsm — openpyxl handles both
            sheets = self._extract_xlsx_sheets(path, is_large)
            # Pivot tables — separate pass (requires read-write mode of openpyxl)
            try:
                pivots_by_sheet = extract_pivots_from_workbook(path)
                if pivots_by_sheet:
                    for s in sheets:
                        if s.sheet_name in pivots_by_sheet:
                            s.pivot_tables = pivots_by_sheet[s.sheet_name]
                            # Force the classification for sheets that contain explicit pivots
                            s.sheet_type = "pivot_summary"
            except Exception as exc:
                print(f"[indexer] pivot scan skipped for {path.name}: {exc}")

        return FileMeta(
            path=path,
            name=path.name,
            file_type=ext.lstrip("."),
            size_mb=round(size_mb, 2),
            modified_at=stat.st_mtime,
            sheets=sheets,
        )

    def _extract_csv_sheet(self, path: Path, is_large: bool) -> SheetMeta:
        # Probe first 15 rows without treating any row as header
        probe = pd.read_csv(path, header=None, nrows=15, low_memory=False, encoding_errors="replace")
        probe_rows = [tuple(probe.iloc[i]) for i in range(len(probe))]
        header_idx = _find_header_row_validated(
            probe_rows,
            read_fn=lambda idx: pd.read_csv(path, header=idx, nrows=3, low_memory=False, encoding_errors="replace"),
        )

        df = pd.read_csv(path, header=header_idx, nrows=200, low_memory=False, encoding_errors="replace")
        headers = [str(c) for c in df.columns.tolist()]
        dtypes = {str(c): str(df[c].dtype) for c in df.columns}
        samples: dict[str, list] = {str(c): _safe_sample(df[c], self.cfg.max_sample_rows) for c in df.columns}

        if is_large:
            try:
                lf = pl.scan_csv(path, skip_rows=header_idx, has_header=True, infer_schema_length=100)
                row_count = lf.select(pl.len()).collect().item()
            except Exception:
                row_count = sum(1 for _ in open(path, encoding="utf-8", errors="ignore")) - header_idx - 1
        else:
            row_count = sum(1 for _ in open(path, encoding="utf-8", errors="ignore")) - header_idx - 1

        return SheetMeta(
            sheet_name="Sheet1",
            row_count=max(0, row_count),
            col_count=len(headers),
            columns=[
                ColumnMeta(name=h, dtype=dtypes.get(h, "unknown"),
                           sample_values=samples.get(h, []))
                for h in headers
            ],
            header_row=header_idx,
            sheet_type=_classify_sheet(headers, max(0, row_count), header_idx),
        )

    def _extract_xlsx_sheets(self, path: Path, is_large: bool) -> list[SheetMeta]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets: list[SheetMeta] = []

        named_range_map: dict[str, list[str]] = {}
        for name, defn in wb.defined_names.items():
            try:
                for sheet_title, _ in defn.destinations:
                    named_range_map.setdefault(sheet_title, []).append(name)
            except Exception:
                pass

        for ws in wb.worksheets:
            try:
                # Collect first 15 rows to detect real header position
                preview: list[tuple] = []
                for row in ws.iter_rows(values_only=True):
                    preview.append(row)
                    if len(preview) >= 15:
                        break

                if not preview:
                    continue

                if is_large:
                    # For large files, validate candidate rows by checking
                    # how many non-empty cells the candidate row has
                    scores = _score_rows(preview[:15])
                    candidates = sorted(range(len(scores)), key=lambda i: -scores[i])
                    header_idx = candidates[0]
                    for cand in candidates[:5]:
                        if scores[cand] <= 0:
                            break
                        row_data = preview[cand]
                        non_empty = sum(1 for v in row_data if not _is_empty_cell(v))
                        total = len(row_data)
                        # Accept if at least 30% of cells are meaningful values
                        if total > 0 and non_empty / total >= 0.3:
                            header_idx = cand
                            break

                    header_row_data = preview[header_idx]
                    headers = [
                        str(h).strip() if not _is_empty_cell(h) else f"col_{i}"
                        for i, h in enumerate(header_row_data)
                    ]
                    headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]
                    data_rows = [r for r in preview[header_idx + 1:] if r]
                    dtypes = _infer_dtypes_from_rows(headers, data_rows)
                    samples: dict[str, list] = {
                        h: [_jsonify(r[i]) for r in data_rows if r and i < len(r) and not _is_empty_cell(r[i])]
                        for i, h in enumerate(headers)
                    }
                    row_count = max(0, (ws.max_row or 0) - header_idx - 1)
                else:
                    # Probe + validated header detection
                    probe = pd.read_excel(
                        path, sheet_name=ws.title, header=None, nrows=15, engine="openpyxl"
                    )
                    probe_rows = [tuple(probe.iloc[i]) for i in range(len(probe))]
                    header_idx = _find_header_row_validated(
                        probe_rows,
                        read_fn=lambda idx: pd.read_excel(
                            path, sheet_name=ws.title, header=idx, nrows=3, engine="openpyxl"
                        ),
                    )
                    try:
                        df = pd.read_excel(
                            path, sheet_name=ws.title,
                            header=header_idx, nrows=200, engine="openpyxl",
                        )
                        dtypes = {str(c): str(df[c].dtype) for c in df.columns}
                        samples = {str(c): _safe_sample(df[c], self.cfg.max_sample_rows)
                                   for c in df.columns}
                        headers = [str(c) for c in df.columns.tolist()]
                    except Exception:
                        dtypes = {}
                        samples = {}
                    row_count = max(0, (ws.max_row or 0) - header_idx - 1)

                sheets.append(SheetMeta(
                    sheet_name=ws.title,
                    row_count=row_count,
                    col_count=len(headers),
                    columns=[
                        ColumnMeta(name=h, dtype=dtypes.get(h, "unknown"),
                                   sample_values=samples.get(h, []))
                        for h in headers
                    ],
                    named_ranges={nr: "" for nr in named_range_map.get(ws.title, [])},
                    header_row=header_idx,
                    sheet_type=_classify_sheet(headers, row_count, header_idx),
                ))
            except Exception as exc:
                print(f"[indexer] sheet '{ws.title}' error: {exc}")

        wb.close()
        return sheets

    def _extract_xls_sheets(self, path: Path) -> list[SheetMeta]:
        """Extract metadata from legacy .xls files using xlrd via pandas."""
        try:
            xl = pd.ExcelFile(path, engine="xlrd")
        except Exception as exc:
            raise ValueError(f"Cannot open .xls file: {exc}") from exc

        sheets: list[SheetMeta] = []
        for sheet_name in xl.sheet_names:
            try:
                probe = xl.parse(sheet_name, header=None, nrows=15)
                if probe.empty:
                    continue
                probe_rows = [tuple(probe.iloc[i]) for i in range(len(probe))]
                header_idx = _find_header_row_validated(
                    probe_rows,
                    read_fn=lambda idx: xl.parse(sheet_name, header=idx, nrows=3),
                )
                df = xl.parse(sheet_name, header=header_idx, nrows=200)
                if df.empty:
                    continue
                headers = [str(c) for c in df.columns.tolist()]
                dtypes = {str(c): str(df[c].dtype) for c in df.columns}
                samples = {str(c): _safe_sample(df[c], self.cfg.max_sample_rows) for c in df.columns}
                try:
                    row_count = xl.parse(sheet_name, header=header_idx, usecols=[0]).shape[0]
                except Exception:
                    row_count = len(df)
                sheets.append(SheetMeta(
                    sheet_name=sheet_name,
                    row_count=row_count,
                    col_count=len(headers),
                    columns=[
                        ColumnMeta(name=h, dtype=dtypes.get(h, "unknown"),
                                   sample_values=samples.get(h, []))
                        for h in headers
                    ],
                    header_row=header_idx,
                    sheet_type=_classify_sheet(headers, row_count, header_idx),
                ))
            except Exception as exc:
                print(f"[indexer] xls sheet '{sheet_name}' error: {exc}")

        return sheets

    # ------------------------------------------------------------------
    # Internal – persistence
    # ------------------------------------------------------------------

    def _persist_file(self, meta: FileMeta) -> None:
        cur = self.db.cursor()
        cur.execute(
            "INSERT OR REPLACE INTO files(path, name, file_type, size_bytes, modified_at, indexed_at) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                str(meta.path),
                meta.name,
                meta.file_type,
                int(meta.size_mb * 1_048_576),
                meta.modified_at,
                time.time(),
            ),
        )
        file_id = cur.lastrowid

        # Register file node in KG
        self.kg.add_file_node(file_id, str(meta.path), meta.name)

        for sheet in meta.sheets:
            headers = [c.name for c in sheet.columns]
            dtypes = {c.name: c.dtype for c in sheet.columns}
            samples = {c.name: [_jsonify(v) for v in c.sample_values] for c in sheet.columns}

            cur.execute(
                "INSERT OR REPLACE INTO sheets "
                "(file_id, sheet_name, row_count, col_count, headers_json, dtypes_json, sample_json, header_row, sheet_type, pivot_json) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    file_id,
                    sheet.sheet_name,
                    sheet.row_count,
                    sheet.col_count,
                    json.dumps(headers),
                    json.dumps(dtypes),
                    json.dumps(samples),
                    sheet.header_row,
                    sheet.sheet_type,
                    json.dumps(sheet.pivot_tables or []),
                ),
            )
            sheet_id = cur.lastrowid

            # Register sheet node in KG
            self.kg.add_sheet_node(
                sheet_id, file_id, str(meta.path),
                sheet.sheet_name, sheet.col_count, sheet.row_count,
            )

            # FTS5 indexing
            for h in headers:
                cur.execute(
                    "INSERT INTO headers_fts(header_name, sheet_id, file_path) VALUES (?, ?, ?)",
                    (h.lower(), sheet_id, str(meta.path)),
                )

        self.db.commit()

    def _delete_file(self, file_id: int) -> None:
        sheet_ids = [
            r[0] for r in self.db.execute(
                "SELECT id FROM sheets WHERE file_id = ?", (file_id,)
            ).fetchall()
        ]
        for sid in sheet_ids:
            self.db.execute("DELETE FROM headers_fts WHERE sheet_id = ?", (sid,))
        self.db.execute("DELETE FROM files WHERE id = ?", (file_id,))
        self.db.commit()

    def _rebuild_kg(self) -> int:
        self.kg = KnowledgeGraph()
        for frow in self.db.execute("SELECT id, path, name FROM files").fetchall():
            self.kg.add_file_node(frow[0], frow[1], frow[2])
        for srow in self.db.execute(
            "SELECT id, file_id, sheet_name, col_count, row_count, sheet_type, "
            "(SELECT path FROM files WHERE id = file_id) FROM sheets"
        ).fetchall():
            self.kg.add_sheet_node(
                srow[0], srow[1], srow[6], srow[2], srow[3], srow[4],
                sheet_type=srow[5] or "data",
            )
        return self.kg.build_column_edges(self.db)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _open_db(self) -> sqlite3.Connection:
        db = sqlite3.connect(str(self.cfg.db_path), check_same_thread=False)
        db.executescript(_DDL)
        db.execute("PRAGMA journal_mode=WAL")
        db.execute("PRAGMA foreign_keys=ON")
        # Schema migrations for old databases
        existing_cols = {r[1] for r in db.execute("PRAGMA table_info(sheets)").fetchall()}
        if "header_row" not in existing_cols:
            db.execute("ALTER TABLE sheets ADD COLUMN header_row INTEGER NOT NULL DEFAULT 0")
        if "sheet_type" not in existing_cols:
            db.execute("ALTER TABLE sheets ADD COLUMN sheet_type TEXT NOT NULL DEFAULT 'data'")
        if "pivot_json" not in existing_cols:
            db.execute("ALTER TABLE sheets ADD COLUMN pivot_json TEXT NOT NULL DEFAULT '[]'")
        db.commit()
        return db

    def _iter_data_files(self):
        for ext in ("*.xlsx", "*.xlsm", "*.xls", "*.csv"):
            yield from self.cfg.data_path.rglob(ext)


# ------------------------------------------------------------------
# Utilities
# ------------------------------------------------------------------

def _safe_sample(series: "pd.Series", n: int) -> list:
    vals = series.dropna().head(n).tolist()
    return [_jsonify(v) for v in vals]


def _jsonify(v):
    if isinstance(v, (int, float, str, bool)) or v is None:
        return v
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, np.floating):
        return float(v)
    return str(v)


_PIVOT_HEADER_SIGNALS = frozenset({
    "grand total", "subtotal", "total", "sum", "average", "count",
    "% of total", "running total",
})
_PIVOT_ROW_SIGNALS = {"grand total", "total", "subtotal"}
_DETAIL_MIN_ROWS = 1_000


def _classify_sheet(headers: list[str], row_count: int, header_row: int) -> str:
    """
    Classify a sheet as 'pivot_summary', 'detail', or 'data'.

    pivot_summary — small row count with aggregation column names
    detail        — high row count, granular transaction-level data
    data          — everything else
    """
    lower = {h.lower().strip() for h in headers}
    has_pivot_col = bool(lower & _PIVOT_HEADER_SIGNALS)
    has_grand_total = "grand total" in lower
    if has_grand_total or (has_pivot_col and row_count < _DETAIL_MIN_ROWS):
        return "pivot_summary"
    if row_count >= _DETAIL_MIN_ROWS:
        return "detail"
    return "data"


def _is_empty_cell(v) -> bool:
    """True if a cell contains no meaningful value (None, NaN, blank string, etc.)."""
    if v is None:
        return True
    # float NaN: NaN != NaN is the canonical Python NaN test
    if isinstance(v, float) and v != v:
        return True
    if isinstance(v, str) and v.strip().lower() in ("", "nan", "none", "null", "n/a", "#n/a"):
        return True
    return False


def _score_rows(rows: list[tuple]) -> list[float]:
    """
    Score each row by how likely it is to be a header row.
    Higher = more string values, fewer numbers = more header-like.
    """
    scores: list[float] = []
    for row in rows:
        if not row:
            scores.append(0.0)
            continue
        non_null = [v for v in row if not _is_empty_cell(v)]
        if not non_null:
            scores.append(0.0)
            continue
        str_count = sum(1 for v in non_null if isinstance(v, str))
        num_count = sum(1 for v in non_null if isinstance(v, (int, float)))
        # Headers = mostly text; data rows = mostly numbers; title/metadata = sparse
        score = str_count * 2.0 - num_count * 0.5 + len(non_null) * 0.2
        scores.append(score)
    return scores


def _find_header_row(rows: list[tuple], max_scan: int = 15) -> int:
    """Return the 0-based index of the most likely header row by score alone."""
    scores = _score_rows(rows[:max_scan])
    if not scores:
        return 0
    return max(range(len(scores)), key=lambda i: scores[i])


def _find_header_row_validated(
    probe_rows: list[tuple],
    read_fn,
    max_candidates: int = 5,
    unnamed_threshold: float = 0.30,
) -> int:
    """
    Find header row with a validation step.

    Tries the top `max_candidates` row indices in score order.
    For each, calls `read_fn(header_idx)` which should return a DataFrame.
    Returns the first candidate whose resulting columns have fewer than
    `unnamed_threshold` fraction of "Unnamed: X" columns.
    Falls back to the highest-scoring candidate if all fail validation.
    """
    scores = _score_rows(probe_rows[:15])
    if not scores:
        return 0

    candidates = sorted(range(len(scores)), key=lambda i: -scores[i])
    best = candidates[0]

    for idx in candidates[:max_candidates]:
        if scores[idx] <= 0:
            break
        try:
            df = read_fn(idx)
            cols = [str(c) for c in df.columns]
            if not cols:
                continue
            unnamed = sum(1 for c in cols if c.startswith("Unnamed:"))
            if unnamed / len(cols) < unnamed_threshold:
                return idx
        except Exception:
            continue

    return best


def _infer_dtypes_from_rows(headers: list[str], rows: list[tuple]) -> dict[str, str]:
    dtypes: dict[str, str] = {}
    for i, h in enumerate(headers):
        vals = [r[i] for r in rows if r and i < len(r) and r[i] is not None]
        if not vals:
            dtypes[h] = "unknown"
        elif all(isinstance(v, bool) for v in vals):
            dtypes[h] = "bool"
        elif all(isinstance(v, int) for v in vals):
            dtypes[h] = "int64"
        elif all(isinstance(v, (int, float)) for v in vals):
            dtypes[h] = "float64"
        else:
            dtypes[h] = "object"
    return dtypes
